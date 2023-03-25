from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime
import time
from selenium.webdriver.common.by import By

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from random import uniform

def get_keyword(text):
    return text.replace(" ", "%20")

def sort_kind(index):
    if index == 1:
        return 'vcount'
    elif index == 2:
        return 'date'
    else:
        return 'none'

def collect_urls(driver, keyword, start_date, end_date, sort_index):
    page_index = 1
    period_txt = f"&period={start_date}.%7C{end_date}."
    _sort_kind = sort_kind(sort_index)
    date = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
    url_list_filename = f"result/url_list_{keyword.replace(' ', '+')}_{date}.txt"

    with open(url_list_filename, 'w') as f:
        page_url = []

        while True:
            time.sleep(uniform(0.01, 1.0))
            url = f'https://kin.naver.com/search/list.nhn?&sort={_sort_kind}&query={get_keyword(keyword)}{period_txt}&section=kin&page={page_index}'
            driver.get(url)
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            tags = soup.find_all('a', class_="_nclicks:kin.txt _searchListTitleAnchor")
            for tag in tags:
                url = tag.get('href').replace('amp;', '')
                page_url.append(url)
                f.write(url + "\n")

            post_number = driver.find_element(By.CLASS_NAME, 'number').text
            post_number = post_number.strip("()")
            current_number = post_number.split('/')[0].split('-')[1].replace(',', '')
            total_number = post_number.split('/')[1].replace(',', '')

            if int(current_number) == int(total_number):
                break
            else:
                page_index += 1

    return page_url, url_list_filename

def extract_data(driver, page_url):
    date = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
    filename = f'result/{keyword.replace(" ", ".")}_{date}_crawling_result.xlsx'

    wb = Workbook()
    sheet = wb.active
    sheet.append(['제목', '질문', '답변'])

    for j in range(1, 4):
        sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

    for i in page_url:
        driver.get(i)
        title = driver.find_element(By.CLASS_NAME, 'title').text

        try:
            question_txt = driver.find_element(By.CLASS_NAME, 'c-heading__content').text
        except:
            question_txt = ""

        answer_list = driver.find_elements(By.CLASS_NAME, "answer-content__list._answerList")

        for n, answer in enumerate(answer_list):
            context_body = answer.find_element(By.CLASS_NAME, "_endContentsText.c-heading-answer__content-user").text
            sheet.append([title, question_txt, context_body])

        wb.save(filename)

    return filename

def main():
    chrome_options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=chrome_options)

    keyword = '의료 심리학'
    driver.get(f'https://kin.naver.com/search/list.nhn?query={get_keyword(keyword)}')
